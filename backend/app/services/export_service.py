"""Build Excel (.xlsx) and PDF exports for opportunities and multi-sheet reports.

Excel uses openpyxl (tables, conditional formatting, charts, hyperlinks); PDF uses
reportlab. openpyxl cannot author *native* interactive Excel PivotTables, so the
"Pivot Summary" sheet provides pre-aggregated category/status pivots + charts.
"""
from __future__ import annotations

import io
from typing import Any, Iterable
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.scoring import score_band

_HEADERS = [
    "ID", "Title", "Country", "Region", "Institution",
    "Category", "Score", "Status", "Budget", "Deadline",
]
_WIDTHS = [38, 48, 16, 22, 28, 18, 8, 16, 16, 14]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
_LINK_FONT = Font(color="0563C1", underline="single")
# Score band fills (spec doc04 §14): >=80 green High, 50-79 amber Medium, <50 grey Low.
_GREEN = PatternFill("solid", fgColor="C6EFCE")
_AMBER = PatternFill("solid", fgColor="FFEB9C")
_GREY = PatternFill("solid", fgColor="E7E6E6")
_BAND_FILL = {"high": _GREEN, "medium": _AMBER, "low": _GREY}


def _enum(value: Any) -> Any:
    return value.value if hasattr(value, "value") else value


def _score_fill(score: Any) -> PatternFill | None:
    if score is None or score == "":
        return None
    return _BAND_FILL[score_band(score)]


def _style_header_row(sheet: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def _write_opportunity_rows(sheet: Worksheet, opportunities: Iterable[Any]) -> None:
    """Write the standard opportunity table (header + rows + formatting + hyperlinks)."""
    sheet.append(_HEADERS)
    _style_header_row(sheet, len(_HEADERS))

    for opp in opportunities:
        deadline = opp.deadline.strftime("%Y-%m-%d") if getattr(opp, "deadline", None) else ""
        score = opp.score if getattr(opp, "score", None) is not None else ""
        sheet.append([
            str(opp.id),
            opp.title,
            getattr(opp, "country", "") or "",
            getattr(opp, "region", "") or "",
            getattr(opp, "institution", "") or "",
            _enum(getattr(opp, "category", "")),
            score,
            _enum(getattr(opp, "status", "")),
            getattr(opp, "budget", "") or "",
            deadline,
        ])
        row = sheet.max_row
        fill = _score_fill(getattr(opp, "score", None))
        if fill is not None:
            sheet.cell(row=row, column=7).fill = fill
        # Hyperlink the Title cell to the source document (Skills/05 §1.3).
        source_url = getattr(opp, "source_url", None)
        if source_url:
            title_cell = sheet.cell(row=row, column=2)
            title_cell.hyperlink = source_url
            title_cell.font = _LINK_FONT

    for index, width in enumerate(_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{sheet.max_row}"


def build_opportunities_workbook(opportunities: Iterable[Any]) -> bytes:
    """Return an .xlsx (bytes) of the given opportunities, formatted per the spec."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Opportunities"
    _write_opportunity_rows(sheet, opportunities)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Multi-sheet intelligence report — Excel (FR-REPORT-004)
# --------------------------------------------------------------------------- #
def _write_opportunity_sheet(workbook: Workbook, title: str, opportunities: Iterable[Any]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    _write_opportunity_rows(sheet, opportunities)


def _write_summary_sheet(sheet: Worksheet, report_type: Any, kpis: dict[str, Any], summary: str) -> None:
    sheet.title = "Executive Summary"
    sheet["A1"] = f"{str(_enum(report_type)).title()} Intelligence Report"
    sheet["A1"].font = _TITLE_FONT

    row = 3
    for label, value in (kpis or {}).items():
        sheet.cell(row=row, column=1, value=str(label)).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Executive Summary").font = Font(bold=True, size=12)
    row += 1
    cell = sheet.cell(row=row, column=1, value=summary or "No summary available.")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row + 14, end_column=6)

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 60


def _add_bar_chart(sheet: Worksheet, title: str, anchor: str, max_row: int) -> None:
    """Bar chart over a 2-column [label, value] table starting at A1 (header in row 1)."""
    if max_row < 2:
        return
    chart = BarChart()
    chart.title = title
    chart.legend = None
    chart.height = 8
    chart.width = 16
    data = Reference(sheet, min_col=2, min_row=1, max_row=max_row)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, anchor)


def _write_keyed_sheet(
    workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]], chart_title: str | None = None
) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.append(headers)
    _style_header_row(sheet, len(headers))
    for row in rows:
        sheet.append(row)
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 28
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    if chart_title and sheet.max_row >= 2:
        _add_bar_chart(sheet, chart_title, f"{get_column_letter(len(headers) + 2)}2", sheet.max_row)


def _write_pivot_sheet(workbook: Workbook, by_category: list[dict[str, Any]], by_status: list[dict[str, Any]]) -> None:
    """A 'Pivot Summary' sheet: pre-aggregated category/status counts + bar charts.

    openpyxl cannot write native interactive PivotTables, so these are equivalent
    pre-computed pivots.
    """
    sheet = workbook.create_sheet(title="Pivot Summary")

    sheet["A1"] = "Category"
    sheet["B1"] = "Opportunities"
    for col in ("A1", "B1"):
        sheet[col].font = _HEADER_FONT
        sheet[col].fill = _HEADER_FILL
    cat_start = 2
    for row in by_category:
        sheet.append([row.get("key") or "Unknown", row.get("count", 0)])
    cat_end = max(cat_start, sheet.max_row)

    status_header = cat_end + 3
    sheet.cell(row=status_header, column=1, value="Status").font = _HEADER_FONT
    sheet.cell(row=status_header, column=1).fill = _HEADER_FILL
    sheet.cell(row=status_header, column=2, value="Opportunities").font = _HEADER_FONT
    sheet.cell(row=status_header, column=2).fill = _HEADER_FILL
    for index, row in enumerate(by_status):
        sheet.cell(row=status_header + 1 + index, column=1, value=row.get("key") or "Unknown")
        sheet.cell(row=status_header + 1 + index, column=2, value=row.get("count", 0))
    status_end = status_header + len(by_status)

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 16

    # Category chart over A1:B{cat_end}
    if cat_end >= 2:
        chart = BarChart()
        chart.title = "Opportunities by Category"
        chart.legend = None
        chart.height = 8
        chart.width = 16
        chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=cat_end), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=cat_end))
        sheet.add_chart(chart, "D2")
    # Status chart
    if status_end >= status_header + 1:
        chart2 = BarChart()
        chart2.title = "Opportunities by Status"
        chart2.legend = None
        chart2.height = 8
        chart2.width = 16
        chart2.add_data(Reference(sheet, min_col=2, min_row=status_header, max_row=status_end), titles_from_data=True)
        chart2.set_categories(Reference(sheet, min_col=1, min_row=status_header + 1, max_row=status_end))
        sheet.add_chart(chart2, "D20")


def build_report_workbook(
    report_type: Any,
    sections: dict[str, Any],
    summary: str,
    kpis: dict[str, Any],
) -> bytes:
    """Return an .xlsx (bytes) for a full intelligence report.

    ``sections`` keys: ``new`` / ``active`` / ``closed`` (Opportunity ORM rows),
    ``regional`` ({region, count, avg_score}), ``standards`` ({standard, mentions}),
    ``by_category`` / ``by_status`` ({key, count}) for the pivot sheet.
    """
    workbook = Workbook()
    _write_summary_sheet(workbook.active, report_type, kpis, summary)

    _write_opportunity_sheet(workbook, "New Opportunities", sections.get("new", []))
    _write_opportunity_sheet(workbook, "Active Opportunities", sections.get("active", []))
    _write_opportunity_sheet(workbook, "Closed Opportunities", sections.get("closed", []))

    _write_keyed_sheet(
        workbook,
        "Regional Summary",
        ["Region", "Opportunities", "Avg Score"],
        [[r.get("region") or "Unknown", r.get("count", 0), r.get("avg_score", 0)] for r in sections.get("regional", [])],
        chart_title="Opportunities by Region",
    )
    _write_keyed_sheet(
        workbook,
        "Standards Summary",
        ["Standard", "Mentions"],
        [[s.get("standard"), s.get("mentions", 0)] for s in sections.get("standards", [])],
        chart_title="Standards Mentions",
    )
    _write_pivot_sheet(workbook, sections.get("by_category", []), sections.get("by_status", []))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Multi-section intelligence report — PDF (FR-REPORT-004)
# --------------------------------------------------------------------------- #
def _pdf_table(headers: list[str], rows: list[list[Any]]) -> Table:
    data = [headers] + [[str(c) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6FB")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def build_report_pdf(
    report_type: Any,
    sections: dict[str, Any],
    summary: str,
    kpis: dict[str, Any],
) -> bytes:
    """Return an executive-shareable PDF (bytes) for the report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, leftMargin=1.6 * cm, rightMargin=1.6 * cm, topMargin=1.6 * cm, bottomMargin=1.6 * cm,
        title=f"{str(_enum(report_type)).title()} Intelligence Report",
    )
    styles = getSampleStyleSheet()
    elements: list[Any] = []

    elements.append(Paragraph(f"{str(_enum(report_type)).title()} Intelligence Report", styles["Title"]))
    elements.append(Spacer(1, 0.3 * cm))

    elements.append(Paragraph("Key Metrics", styles["Heading2"]))
    elements.append(_pdf_table(["Metric", "Value"], [[k, v] for k, v in (kpis or {}).items()]))
    elements.append(Spacer(1, 0.4 * cm))

    elements.append(Paragraph("Executive Summary", styles["Heading2"]))
    summary_html = escape(summary or "No summary available.").replace("\n", "<br/>")
    elements.append(Paragraph(summary_html, styles["BodyText"]))
    elements.append(Spacer(1, 0.4 * cm))

    top = (sections.get("active") or [])[:15]
    if top:
        elements.append(Paragraph("Top Active Opportunities", styles["Heading2"]))
        rows = [[(o.title or "")[:48], o.region or "", o.score if o.score is not None else "", _enum(o.status)] for o in top]
        elements.append(_pdf_table(["Title", "Region", "Score", "Status"], rows))
        elements.append(Spacer(1, 0.4 * cm))

    regional = sections.get("regional") or []
    if regional:
        elements.append(Paragraph("Regional Summary", styles["Heading2"]))
        rows = [[r.get("region") or "Unknown", r.get("count", 0), r.get("avg_score", 0)] for r in regional]
        elements.append(_pdf_table(["Region", "Opportunities", "Avg Score"], rows))

    doc.build(elements)
    return buffer.getvalue()
